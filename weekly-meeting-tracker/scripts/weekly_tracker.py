#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
周会需求进度追踪脚本

用法:
    python weekly_tracker.py <excel_path> [--output <output_path>]

功能:
    1. 读取多 sheet 的 Excel 文件（sheet 名为 4 位日期如 0821）
    2. 解析上期「跟进记录」中的下一节点时间和预期动作
    3. 在本期 sheet 中比对状态变更
    4. 生成「AI进度判断」列并写入 Excel
    5. 输出汇总统计

依赖: openpyxl, pandas
"""

import re
import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================
# 常量定义
# ============================================================

EXPECTED_HEADERS = [
    '编号', '系统模块', '跟进记录', '需求名称', '状态',
    'IT标签', '实际完时', '业务标签', '耗时(天)', '第一提出人',
    '提出日期', '跟进意见', '任务工时(h)', '业务跟进人', 'IT跟进人',
    '类型', '紧急度', '预计完时'
]

# 状态流转顺序（数字越大越靠后）
STATUS_ORDER = {
    '待IT跟进': 1,
    '已IT跟进': 2,
    '设计中': 3,
    '设计完成': 4,
    '待研发': 5,
    '研发中': 6,
    '待上线': 7,
    '待验收': 8,
}

# 关键词 → 期望状态映射（值为列表，满足其一即可）
ACTION_TO_STATUS = {
    '评审': ['设计完成', '待研发'],
    '需求评审': ['设计完成', '待研发'],
    '技术评审': ['设计完成', '待研发'],
    '方案评审': ['设计完成', '待研发'],
    '完成设计': ['设计完成'],
    '出方案': ['设计完成'],
    '定方案': ['设计完成'],
    '输出方案': ['设计完成'],
    '出草稿': ['设计完成'],
    '设计': ['设计完成'],
    '研发': ['研发中'],
    '开发': ['研发中'],
    '进入开发': ['研发中'],
    '排期开发': ['研发中'],
    '进入研发': ['研发中'],
    '介入开发': ['研发中'],
    '上线': ['待上线', '待验收'],
    '发布': ['待上线', '待验收'],
    '验收': ['待验收'],
    '提测': ['待上线'],
    '测试': ['待上线'],
    '联调': ['待上线'],
    '需求沟通': ['已IT跟进', '设计中'],
    '需求对接': ['已IT跟进', '设计中'],
    '需求澄清': ['已IT跟进', '设计中'],
    '对接需求': ['已IT跟进', '设计中'],
    '沟通需求': ['已IT跟进', '设计中'],
    '打标': ['已IT跟进', '设计中'],
    '业务打标': ['已IT跟进', '设计中'],
}

# 动作关键词 → 人话描述
ACTION_HUMAN_MAP = {
    '评审': '完成评审',
    '需求评审': '完成需求评审',
    '技术评审': '完成技术评审',
    '方案评审': '完成方案评审',
    '完成设计': '完成设计',
    '出方案': '出方案',
    '定方案': '确定方案',
    '输出方案': '输出方案',
    '出草稿': '出草稿',
    '设计': '完成设计',
    '研发': '进入研发',
    '开发': '进入开发',
    '进入开发': '进入开发',
    '排期开发': '排期进入开发',
    '进入研发': '进入研发',
    '介入开发': '开发介入',
    '上线': '完成上线',
    '发布': '发布上线',
    '验收': '完成验收',
    '提测': '提交测试',
    '测试': '进入测试',
    '联调': '完成联调',
    '需求沟通': '完成需求沟通',
    '需求对接': '完成需求对接',
    '需求澄清': '完成需求澄清',
    '对接需求': '完成需求对接',
    '沟通需求': '完成需求沟通',
    '打标': '完成业务打标',
    '业务打标': '完成业务打标',
}

# 星期映射
WEEKDAY_MAP = {
    '一': 0, '二': 1, '三': 2, '四': 3,
    '五': 4, '六': 5, '日': 6, '天': 6
}


# ============================================================
# 工具函数
# ============================================================

def clean_html(text):
    """清除 HTML 标签"""
    if not text or not isinstance(text, str):
        return text or ''
    return re.sub(r'<[^>]+>', '', text).strip()


def parse_sheet_date(sheet_name, year=None):
    """将 sheet 名称（如 '0821'）解析为 datetime 对象。"""
    if year is None:
        year = datetime.now().year
    match = re.match(r'^(\d{2})(\d{2})$', sheet_name.strip())
    if not match:
        return None
    month, day = int(match.group(1)), int(match.group(2))
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def _extract_follow_date(text, sheet_date):
    """提取跟进记录开头的跟进日期（如 '0820：...' 中的 0820）。"""
    m = re.match(r'^(\d{2})(\d{2})\s*[：:\s]', text)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        try:
            return datetime(sheet_date.year, month, day)
        except ValueError:
            pass
    return sheet_date


def _strip_leading_date(text):
    """去掉跟进记录开头的跟进日期前缀。"""
    cleaned = re.sub(r'^\d{4}\s*[：:\s]\s*', '', text)
    return cleaned


def extract_time_nodes(text, sheet_date):
    """
    从跟进记录文本中提取时间节点和预期动作。
    返回: (node_date: datetime or None, action_keywords: list[str], raw_match: str)
    """
    if not text or not isinstance(text, str):
        return None, [], ''

    text = clean_html(text)
    if not text:
        return None, [], ''

    year = sheet_date.year
    sheet_month = sheet_date.month
    body = _strip_leading_date(text)
    candidates = []

    # --- 明确日期格式 ---
    for m in re.finditer(r'(\d{1,2})[/\-.](\d{1,2})', body):
        month, day = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                dt = datetime(year, month, day)
                if (sheet_month - month) > 6:
                    dt = datetime(year + 1, month, day)
                elif (month - sheet_month) > 6:
                    dt = datetime(year - 1, month, day)
                context = body[m.start():min(m.end() + 20, len(body))]
                candidates.append((dt, context, m.end()))
            except ValueError:
                pass

    # MMDD 4位数字格式
    for m in re.finditer(r'(\d{4})(?=[\u4e00-\u9fff])', body):
        digits = m.group(1)
        month, day = int(digits[:2]), int(digits[2:])
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                dt = datetime(year, month, day)
                if (sheet_month - month) > 6:
                    dt = datetime(year + 1, month, day)
                context = body[m.start():min(m.end() + 20, len(body))]
                candidates.append((dt, context, m.end()))
            except ValueError:
                pass

    # "XX日" 或 "XX号" 格式
    for m in re.finditer(r'(\d{1,2})[日号]', body):
        day = int(m.group(1))
        if 1 <= day <= 31:
            month = sheet_month
            try:
                dt = datetime(year, month, day)
                if dt < sheet_date - timedelta(days=14):
                    month += 1
                    if month > 12:
                        month = 1
                        year += 1
                    dt = datetime(year, month, day)
                context = body[max(0, m.start() - 10):min(m.end() + 20, len(body))]
                candidates.append((dt, context, m.end()))
            except ValueError:
                pass

    # "X月XX" 格式
    for m in re.finditer(r'(\d{1,2})月(\d{1,2})[日号]?', body):
        month, day = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                dt = datetime(year, month, day)
                if (sheet_month - month) > 6:
                    dt = datetime(year + 1, month, day)
                context = body[m.start():min(m.end() + 20, len(body))]
                candidates.append((dt, context, m.end()))
            except ValueError:
                pass

    # --- 相对时间格式 ---
    for m in re.finditer(r'下周([一二三四五六日天])', body):
        wd = WEEKDAY_MAP.get(m.group(1))
        if wd is not None:
            days_ahead = (7 - sheet_date.weekday()) + wd
            dt = sheet_date + timedelta(days=days_ahead)
            context = body[m.start():min(m.end() + 20, len(body))]
            candidates.append((dt, context, m.end()))

    for m in re.finditer(r'本周([一二三四五六日天])', body):
        wd = WEEKDAY_MAP.get(m.group(1))
        if wd is not None:
            days_diff = wd - sheet_date.weekday()
            dt = sheet_date + timedelta(days=days_diff)
            context = body[m.start():min(m.end() + 20, len(body))]
            candidates.append((dt, context, m.end()))

    for m in re.finditer(r'明天|明日', body):
        follow_date = _extract_follow_date(text, sheet_date)
        dt = follow_date + timedelta(days=1)
        context = body[m.start():min(m.end() + 20, len(body))]
        candidates.append((dt, context, m.end()))

    for m in re.finditer(r'后天', body):
        follow_date = _extract_follow_date(text, sheet_date)
        dt = follow_date + timedelta(days=2)
        context = body[m.start():min(m.end() + 20, len(body))]
        candidates.append((dt, context, m.end()))

    for m in re.finditer(r'今天|今日', body):
        follow_date = _extract_follow_date(text, sheet_date)
        dt = follow_date
        context = body[m.start():min(m.end() + 20, len(body))]
        candidates.append((dt, context, m.end()))

    for m in re.finditer(r'下周(?![一二三四五六日天])', body):
        days_ahead = (7 - sheet_date.weekday()) + 2
        dt = sheet_date + timedelta(days=days_ahead)
        context = body[m.start():min(m.end() + 20, len(body))]
        candidates.append((dt, context, m.end()))

    for m in re.finditer(r'(?<![下本])周([一二三四五六日天])', body):
        wd = WEEKDAY_MAP.get(m.group(1))
        if wd is not None:
            follow_date = _extract_follow_date(text, sheet_date)
            days_diff = wd - follow_date.weekday()
            if days_diff <= 0:
                days_diff += 7
            dt = follow_date + timedelta(days=days_diff)
            context = body[m.start():min(m.end() + 20, len(body))]
            candidates.append((dt, context, m.end()))

    if not candidates:
        return None, [], ''

    candidates.sort(key=lambda x: x[0], reverse=True)
    best_date, best_context, _ = candidates[0]
    actions = _extract_actions(text)

    return best_date, actions, best_context


def _extract_actions(text):
    """从文本中提取所有匹配的动作关键词"""
    text = clean_html(text)
    matched = []
    for kw in sorted(ACTION_TO_STATUS.keys(), key=len, reverse=True):
        if kw in text:
            matched.append(kw)
    return matched


def get_expected_statuses(actions):
    """根据动作关键词列表，获取所有期望状态"""
    statuses = set()
    for action in actions:
        if action in ACTION_TO_STATUS:
            statuses.update(ACTION_TO_STATUS[action])
    return list(statuses)


def action_to_human(actions):
    """把动作关键词翻译成用户能看懂的描述"""
    if not actions:
        return ''
    primary = actions[0]
    return ACTION_HUMAN_MAP.get(primary, primary)


def expected_status_human(expected_statuses):
    """取期望状态中最低要求的那个"""
    if not expected_statuses:
        return ''
    min_order = min(STATUS_ORDER.get(s, 99) for s in expected_statuses)
    for s in expected_statuses:
        if STATUS_ORDER.get(s, 99) == min_order:
            return s
    return expected_statuses[0]


def status_reached(current_status, expected_statuses):
    """判断当前状态是否达到或超过任一期望状态。"""
    if not expected_statuses:
        return 'unknown'
    current_order = STATUS_ORDER.get(current_status, 0)
    expected_orders = [STATUS_ORDER.get(s, 0) for s in expected_statuses]
    min_expected = min(expected_orders)
    if current_order >= min_expected:
        return 'reached'
    elif current_order > 0:
        return 'progressed'
    else:
        return 'not_reached'


# ============================================================
# 主流程
# ============================================================

def load_and_validate(excel_path):
    """Step 1: 加载并校验 Excel。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_names = wb.sheetnames
    warnings = []

    date_sheets = []
    for name in sheet_names:
        dt = parse_sheet_date(name)
        if dt:
            date_sheets.append((name, dt))

    if not date_sheets:
        warnings.append(f"未找到 4 位日期格式的 sheet 名称。现有 sheet: {sheet_names}")
        return wb, [], warnings

    date_sheets.sort(key=lambda x: x[1])

    last_sheet = wb[date_sheets[-1][0]]
    actual_headers = [last_sheet.cell(1, c).value for c in range(1, last_sheet.max_column + 1)]
    actual_18 = actual_headers[:18]
    if actual_18 != EXPECTED_HEADERS:
        mismatches = []
        for i, (a, e) in enumerate(zip(actual_18, EXPECTED_HEADERS)):
            if a != e:
                mismatches.append(f"列{i+1}: 期望'{e}', 实际'{a}'")
        if mismatches:
            warnings.append(f"表头不完全匹配: {'; '.join(mismatches)}")

    return wb, date_sheets, warnings


def read_sheet_data(ws):
    """读取一个 sheet 的数据为字典，key 为编号"""
    data = {}
    for row in range(2, ws.max_row + 1):
        req_id = ws.cell(row, 1).value
        if req_id is None:
            continue
        try:
            req_id = int(float(req_id))
        except (ValueError, TypeError):
            continue
        record = {
            '编号': req_id,
            '系统模块': ws.cell(row, 2).value or '',
            '跟进记录': ws.cell(row, 3).value or '',
            '需求名称': ws.cell(row, 4).value or '',
            '状态': ws.cell(row, 5).value or '',
            'IT标签': ws.cell(row, 6).value or '',
            '实际完时': ws.cell(row, 7).value,
            '业务标签': ws.cell(row, 8).value or '',
            '耗时(天)': ws.cell(row, 9).value,
            '第一提出人': ws.cell(row, 10).value or '',
            '提出日期': ws.cell(row, 11).value,
            '跟进意见': ws.cell(row, 12).value or '',
            '任务工时(h)': ws.cell(row, 13).value,
            '业务跟进人': ws.cell(row, 14).value or '',
            'IT跟进人': ws.cell(row, 15).value or '',
            '类型': ws.cell(row, 16).value or '',
            '紧急度': ws.cell(row, 17).value or '',
            '预计完时': ws.cell(row, 18).value,
            '_row': row,
        }
        data[req_id] = record
    return data


def generate_first_period(ws, sheet_date, sheet_name):
    """
    首期分析：解析每条需求的跟进记录，生成「下期验证要点」列。
    文案风格：直接告诉读者"这条需求计划什么时候做什么，届时应该看到什么状态"。
    """
    ai_col = ws.max_column + 1

    header_cell = ws.cell(1, ai_col)
    header_cell.value = '下期验证要点'
    header_cell.font = Font(bold=True, color='FFFFFF')
    header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions[openpyxl.utils.get_column_letter(ai_col)].width = 55

    stats = {'有明确计划': 0, '无明确计划': 0, '总需求数': 0}

    for row in range(2, ws.max_row + 1):
        req_id = ws.cell(row, 1).value
        if req_id is None:
            continue
        try:
            req_id = int(float(req_id))
        except (ValueError, TypeError):
            continue

        stats['总需求数'] += 1
        follow_text = ws.cell(row, 3).value or ''
        status = ws.cell(row, 5).value or ''

        node_date, actions, _ = extract_time_nodes(follow_text, sheet_date)
        expected_statuses = get_expected_statuses(actions) if actions else []

        cell = ws.cell(row, ai_col)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        if node_date and actions:
            stats['有明确计划'] += 1
            node_str = f'{node_date.month}/{node_date.day}'
            action_human = action_to_human(actions)
            expected_human = expected_status_human(expected_statuses)
            cell.value = f'计划 {node_str} {action_human}，届时状态应变为「{expected_human}」或更后阶段'
            cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

        elif node_date and not actions:
            stats['有明确计划'] += 1
            node_str = f'{node_date.month}/{node_date.day}'
            cell.value = f'提及时间节点 {node_str}，但未明确下一步动作，建议补充'
            cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

        elif not node_date and actions:
            stats['无明确计划'] += 1
            action_human = action_to_human(actions)
            cell.value = f'提及将要{action_human}，但未写明具体时间，建议补充日期'
            cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

        else:
            stats['无明确计划'] += 1
            cell.value = '未提及下一步计划和时间节点'
            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

    return stats


def generate_ai_judgment(prev_data, curr_data, prev_sheet_name, prev_sheet_date, curr_sheet_date):
    """
    跨周比对：生成 AI 判断。
    文案风格：直接说结论 + 原因，让人一眼看懂。
    """
    judgments = {}
    stats = {
        '按期推进': 0,
        '未按期推进': 0,
        '有进展未达预期': 0,
        '需求已移出': 0,
        '节点未到': 0,
        '无明确节点': 0,
        '新增需求': 0,
        '本期需求总数': len(curr_data),
        '有明确节点的需求': 0,
    }
    delayed_list = []

    # 新增需求
    for req_id in curr_data:
        if req_id not in prev_data:
            stats['新增需求'] += 1
            judgments[req_id] = '本期新增需求，无历史数据可比对'

    # 遍历上期需求
    for req_id, prev_record in prev_data.items():
        follow_text = prev_record.get('跟进记录', '')
        prev_status = prev_record.get('状态', '')

        node_date, actions, _ = extract_time_nodes(follow_text, prev_sheet_date)

        if node_date is None or not actions:
            stats['无明确节点'] += 1
            if req_id in curr_data:
                curr_status = curr_data[req_id].get('状态', '')
                if curr_status != prev_status:
                    judgments[req_id] = (
                        f'上期无明确计划，但状态从「{prev_status}」变为「{curr_status}」，有推进'
                    )
                else:
                    judgments[req_id] = (
                        f'上期无明确计划，状态仍为「{curr_status}」，无变化'
                    )
            else:
                stats['需求已移出'] += 1
                stats['无明确节点'] -= 1
                judgments[req_id] = f'本期未找到该需求（编号 {req_id}），可能已关闭或转移'
            continue

        stats['有明确节点的需求'] += 1
        expected_statuses = get_expected_statuses(actions)
        node_str = f'{node_date.month}/{node_date.day}'
        action_human = action_to_human(actions)
        expected_human = expected_status_human(expected_statuses)

        # 本期找不到
        if req_id not in curr_data:
            stats['需求已移出'] += 1
            judgments[req_id] = f'本期未找到该需求（编号 {req_id}），可能已关闭或转移'
            continue

        curr_status = curr_data[req_id].get('状态', '')

        # 节点未到
        if node_date > curr_sheet_date:
            stats['节点未到'] += 1
            judgments[req_id] = (
                f'上期计划 {node_str} {action_human}，时间还没到，下期再看'
            )
            continue

        # 比对状态
        result = status_reached(curr_status, expected_statuses)
        delay_days = (curr_sheet_date - node_date).days

        if result == 'reached':
            stats['按期推进'] += 1
            judgments[req_id] = (
                f'✅ 已按计划推进｜上期计划 {node_str} {action_human}，'
                f'当前状态「{curr_status}」，符合预期'
            )
        elif curr_status == prev_status:
            stats['未按期推进'] += 1
            judgments[req_id] = (
                f'❌ 未按计划推进｜上期计划 {node_str} {action_human}，'
                f'当前状态仍为「{curr_status}」，已延期 {delay_days} 天'
            )
            delayed_list.append({
                '编号': req_id,
                '需求名称': prev_record.get('需求名称', ''),
                '当前状态': curr_status,
                '预期动作': action_human,
                '预期日期': node_str,
                '延期天数': delay_days,
                'IT跟进人': curr_data[req_id].get('IT跟进人', ''),
            })
        else:
            stats['有进展未达预期'] += 1
            judgments[req_id] = (
                f'⚠️ 有进展但未达预期｜上期计划 {node_str} {action_human}（期望「{expected_human}」），'
                f'当前状态为「{curr_status}」，有推进但还没到位'
            )

    return judgments, stats, delayed_list


def write_output(excel_path, wb, curr_sheet_name, curr_data, judgments, output_path=None):
    """将 AI 判断列写入 Excel 并保存。"""
    ws = wb[curr_sheet_name]
    ai_col = ws.max_column + 1

    header_cell = ws.cell(1, ai_col)
    header_cell.value = 'AI进度判断'
    header_cell.font = Font(bold=True, color='FFFFFF')
    header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions[openpyxl.utils.get_column_letter(ai_col)].width = 60

    for row in range(2, ws.max_row + 1):
        req_id = ws.cell(row, 1).value
        if req_id is not None:
            try:
                req_id = int(float(req_id))
            except (ValueError, TypeError):
                continue
            judgment = judgments.get(req_id, '')
            cell = ws.cell(row, ai_col)
            cell.value = judgment
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if '✅' in judgment:
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            elif '❌' in judgment:
                cell.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
                cell.font = Font(color='C62828')
            elif '⚠️' in judgment:
                cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    if output_path is None:
        stem = Path(excel_path).stem
        output_path = str(Path(excel_path).parent / f'{stem}_AI分析结果.xlsx')

    wb.save(output_path)
    return output_path


def print_summary(stats, delayed_list, prev_sheet_name, curr_sheet_name):
    """打印汇总统计"""
    print('\n' + '=' * 60)
    print(f'📊 周会进度追踪（{prev_sheet_name} → {curr_sheet_name}）')
    print('=' * 60)

    print(f'\n本期需求总数: {stats["本期需求总数"]}')
    print(f'有明确计划的需求: {stats["有明确节点的需求"]}')
    print()
    print(f'  ✅ 已按计划推进: {stats["按期推进"]}')
    print(f'  ❌ 未按计划推进: {stats["未按期推进"]}')
    print(f'  ⚠️  有进展未达预期: {stats["有进展未达预期"]}')
    print(f'  🔴 需求已移出: {stats["需求已移出"]}')
    print(f'  ⏳ 时间还没到: {stats["节点未到"]}')
    print(f'  ℹ️  上期无明确计划: {stats["无明确节点"]}')
    print(f'  🆕 本期新增: {stats["新增需求"]}')

    if delayed_list:
        print('\n' + '-' * 60)
        print('🚨 未按计划推进的需求（建议周会重点讨论）：')
        print('-' * 60)
        for item in delayed_list:
            print(
                f'  [{item["编号"]}] {item["需求名称"][:30]}'
                f'  | 当前: {item["当前状态"]}'
                f'  | 计划: {item["预期日期"]} {item["预期动作"]}'
                f'  | 延期: {item["延期天数"]}天'
                f'  | IT: {item["IT跟进人"]}'
            )

    print('\n' + '=' * 60)


def main():
    parser = argparse.ArgumentParser(description='周会需求进度追踪分析')
    parser.add_argument('excel_path', help='Excel 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径')
    args = parser.parse_args()

    excel_path = args.excel_path
    output_path = args.output

    print(f'📂 加载文件: {excel_path}')

    wb, date_sheets, warnings = load_and_validate(excel_path)

    for w in warnings:
        print(f'⚠️  {w}')

    if len(date_sheets) < 1:
        print('❌ 未找到有效的日期 sheet，请检查 sheet 命名格式（如 0821）。')
        return

    if len(date_sheets) == 1:
        # 首期：只做基线解析
        sheet_name, sheet_date = date_sheets[0]
        print(f'📅 首期数据: {sheet_name} ({sheet_date.strftime("%Y-%m-%d")})')
        print('ℹ️  仅有 1 个 sheet，生成首期基线分析（下期验证要点）。')

        ws = wb[sheet_name]
        stats = generate_first_period(ws, sheet_date, sheet_name)

        if output_path is None:
            stem = Path(excel_path).stem
            output_path = str(Path(excel_path).parent / f'{stem}_首期分析.xlsx')
        wb.save(output_path)

        print(f'\n💾 结果已保存: {output_path}')
        print(f'\n总需求数: {stats["总需求数"]}')
        print(f'有明确计划: {stats["有明确计划"]}')
        print(f'无明确计划: {stats["无明确计划"]}')
        return

    # 多期：跨周比对
    prev_sheet_name, prev_sheet_date = date_sheets[-2]
    curr_sheet_name, curr_sheet_date = date_sheets[-1]

    print(f'📅 上期: {prev_sheet_name} ({prev_sheet_date.strftime("%Y-%m-%d")})')
    print(f'📅 本期: {curr_sheet_name} ({curr_sheet_date.strftime("%Y-%m-%d")})')

    prev_data = read_sheet_data(wb[prev_sheet_name])
    curr_data = read_sheet_data(wb[curr_sheet_name])

    print(f'📋 上期需求数: {len(prev_data)}，本期需求数: {len(curr_data)}')

    judgments, stats, delayed_list = generate_ai_judgment(
        prev_data, curr_data, prev_sheet_name, prev_sheet_date, curr_sheet_date
    )

    output_file = write_output(excel_path, wb, curr_sheet_name, curr_data, judgments, output_path)
    print(f'\n💾 结果已保存: {output_file}')

    print_summary(stats, delayed_list, prev_sheet_name, curr_sheet_name)


if __name__ == '__main__':
    main()
